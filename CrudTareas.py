from datetime import datetime
from openpyxl import load_workbook

Rut="C:\\Users\\SENA\\Desktop\\haroldway\\repositorio_actividad\\Crud_tareas_y_productos.xlsx"
Rut=r"C:\Users\SENA\Desktop\haroldway\repositorio_actividad\Crud_tareas_y_productos.xlsx"

def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Tarea']
    Hoja_datos=Hoja_datos['A2':'F'+ str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
            'estado':i[3].value, 'fecha de inicio':i[4].value,
            'fecha de finalizacion':i[5].value})

    if not(extraer=='todo'):
        info=filtrar(info,extraer )

    for i in info:
        print('********Tarea********')
        print('Id:'+str(i)+'\n'+'Titulo:'+str(info[i]['tarea'])+'\n'+'Descripcion:'+ str(info[i]['descripcion'])+
        '\n'+'Estado:'+str(info[i]['estado'])
        + '\n'+'Fecha Creacion:'+str(info[i]['fecha de inicio'])+
        '\n'+'fecha de finalizacion:'+str(info[i]['fecha de finalizacion']))
        print()
    return

def filtrar(info:dict,filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux
