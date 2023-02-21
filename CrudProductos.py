from datetime import datetime
from openpyxl import load_workbook

Rut="C:\\Users\\SENA\\Desktop\\haroldway\\repositorio_actividad\\Crud_tareas_y_productos.xlsx"
Rut=r"C:\Users\SENA\Desktop\haroldway\repositorio_actividad\Crud_tareas_y_productos.xlsx"

def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Producto']
    Hoja_datos=Hoja_datos['A2':'F'+ str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'Producto':i[1].value, 'Categoria':i[2].value,
            'Precio Unidad':i[3].value, 'Cantidad':i[4].value})

    if not(extraer=='todo'):
        info=filtrar(info,extraer )

    for i in info:
        print('***Producto***')
        print('Id:'+str(i)+'\n'+'Titulo:'+str(info[i]['Producto'])+'\n'+'Categoria:'+ str(info[i]['categoria'])+
        '\n'+'Precio Unidad:'+str(info[i]['precio unidad'])
        + '\n'+'Cantidad:'+str(info[i]['cantidad']))
        print()
    return
 
def filtrar(info:dict,filtro:str):
    aux={}
    for i in info:
        if info[i]['Categoria']==filtro:
            aux.setdefault(i,info[i])
    return aux
